# coding=utf-8
# Version:Python 3.8.0
# Tools:PyCharm 2020.2 x64
__date__ = '2021/5/14 14:26'
__author__ = '张勇'
# 读取excel
# import xlrd
# 读取或 创建excel
import openpyxl
# 读
# book = xlrd.open_workbook('d.xls')  # 获取Excel
#
# print(book.nsheets)  ## 获取sheet数量
# print(book.sheet_names())  #获取所有sheets 名字
# # book.sheets() # 获取所有表单
#
# s1 = book.sheet_by_index(0)  # 通过索引获取sheet
#
#
# print(s1.nrows,s1.name,s1.ncols,s1.number)
#
# # 获取单元格内容
# print(s1.cell_value(rowx=2,colx=3))
#
# # 获取横排内容
# print('第一行的数据为：',s1.row_values(rowx=0))
#
# # 获取列内容
# ages = s1.col_values(colx=2,start_rowx=1)
# print(f'年龄列表为：{ages}')
# print(f'第3列的年龄总和为：{sum(ages)}')
#===============================================

# 创建excel
book = openpyxl.Workbook()
# 当前sheet
sh = book.active
sh.title = '商品列表'
# 單元格寫入数据
# sh['A1'] = 'id'
# sh['B1'] = 'name'
# print(sh['B1'].value)
# # 新建sheet表单，放在最后
# sh1 = book.create_sheet('工资表')
# # 新建sheet表单，放在最前
# sh2 = book.create_sheet('时间表',0)
name2age = {
    "zhangsan":19,
    "zhangsan2":19,
    "zhangsan3":19,
    "zhangsan3":39,
    "zhangsan4":192,
}
# 2单元格写入数据
# sh.cell(2,2,value='hello')
# sh.cell(2,3).value = 'world'
#2.读取单元格内容
# print(sh.cell(2,2).value)

sh['A1'] = '姓名'
sh['B1'] = '年龄'
row = 2
for name,age in name2age.items():
    sh.cell(row,1).value = name
    sh.cell(row,2).value = age
    row+=1
book.save('goods.xlsx')