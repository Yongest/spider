# coding=utf-8
# Version:Python 3.8.0
# Tools:PyCharm 2020.2 x64
__date__ = '2021/5/14 14:26'
__author__ = '张勇'
import requests
import openpyxl
import json
import jsonpath
import time
domain = 'http://127.0.0.1:8080/data/'

# 创建excel
book = openpyxl.Workbook()
# 当前sheet
sh = book.active
sh.title = '商品列表'

name_list = ["商品名字", '品牌', '材质', '市场价', '编码', '型号']
for index0, item in enumerate(name_list):
    sh.cell(1, index0 + 1).value = item

row = 2
for index in range(1,113):
    print('第',index,'页')

    response = requests.get(domain+'new_goods_'+str(index)+'.json')
    json_item = json.loads(response.content.decode())

    # 循环列表
    for item in json_item:
        # 获取材质
        material = ''
        for item2 in item['params']:
            for li in item2['list']:
                if li['name']=='材质':
                    material = li['value']

        # goods_list = json.dumps(item['goods_list'],ensure_ascii=False)
        # print(row)
        priceList = jsonpath.jsonpath(item['goods_list'],'$..market_price')
        # print(priceList, 111)

        codeList = jsonpath.jsonpath(item['goods_list'],'$..code')
        modelList = jsonpath.jsonpath(item['goods_list'],'$..model')
        # print(codeList,priceList,modelList)
        if priceList and len(priceList):
            for goods_index,goods_item in enumerate(priceList):
                sh.cell(row, 1).value = item['name']
                sh.cell(row, 2).value = item['brand']
                sh.cell(row, 3).value = material
                sh.cell(row, 4).value = priceList[goods_index]
                sh.cell(row, 5).value = codeList[goods_index]
                sh.cell(row, 6).value = modelList[goods_index]
                row += 1

    time.sleep(0.3)

book.save('./data/goods.xlsx')